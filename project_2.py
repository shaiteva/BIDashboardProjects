from tkinter.ttk import *
from tkinter import *
import pandas as pd
from matplotlib import pyplot as plt
import glob
from openpyxl.reader.excel import load_workbook

#פונקציה ראשית לבניית חלונות נוספים שמופעלת כאשר לוחצים
# על send בחלון הראשון.יש שימוש בהתניות כדי שהפונקציה תציג
# את החלון הבא בלי שהקוד יקרוס
def check_val():
    if len(text1.get()) == 0:
        c_lbl = Label(window, text="  Please enter a shop name  ", fg="red")
        c_lbl.grid(row=25, column=1)
    else:
        df1 = pd.read_excel('report_2019.xlsx')
        df1.to_csv('report_2019.csv', index=False, quotechar="'")
        df = pd.read_csv('report_2019.csv')

        #התניות בעת כניסה לדוח היומי
        if chek_btn1.get() == 1 and text1.get() != df.columns[0]:
            get_name = Label(window, text="  wrong shop try again    ", fg="black")
            get_name.grid(row=25, column=1)

        elif chek_btn1.get() == 1 and text1.get() == df.columns[0]:
            window.destroy()

            #פונקציה המציגה דוח יומי
            #משתמש בפנדס ליצירת פעולות שונות שיוצגו בתור נתונים על הגרף ולבסוף יוצר את הייצוג הגרפי בתוך חלון
            def chek_date():
                global enumerate
                df1 = 'report_2019.xlsx'
                files = glob.iglob(df1)
                for fileName in files:
                    wb = load_workbook(fileName)
                    w = wb.sheetnames
                    for sheetname in w:
                        try:
                            if sheetname != combo1.get():
                                get_empty = Label(root1, text="", fg="red")
                                get_empty.grid(column=4, row=40)

                            else:
                                df1 = pd.read_excel('report_2019.xlsx', sheet_name=combo1.get())
                                df1.to_csv('report_2019.csv', index=False, quotechar="'")
                                df = pd.read_csv('report_2019.csv')
                                columns = df.columns[3:]

                                for i in columns:
                                    if i == ent_text.get():

                                        df1 = pd.read_excel('report_2019.xlsx', sheet_name=combo1.get())
                                        df1.to_csv('report_2019.csv', index=False, quotechar="'")
                                        df = pd.read_csv('report_2019.csv')

                                        u = df[ent_text.get()]
                                        o = df["costumer price"]
                                        k = df["purchase"]
                                        Employee = 400
                                        Revenue = u * o
                                        Expenses = u * k
                                        Profits_b = Revenue - Expenses

                                        tax=0.10
                                        Profits_b_sum=Profits_b.sum(axis=0, skipna=True)
                                        net_income=Profits_b.sum(axis=0, skipna=True)-(Profits_b.sum(axis=0, skipna=True)*tax)
                                        after_pay_employee = Profits_b.sum(axis=0, skipna=True) - (Profits_b.sum(axis=0, skipna=True)*tax) - Employee

                                        df.set_index('gali', inplace=True)
                                        labels = df.index
                                        sizes = Profits_b
                                        x_axis = df.index
                                        y_axis = df.loc[:][ent_text.get()]
                                        root1.destroy()

                                        fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=1)
                                        fig.canvas.set_window_title('daily report 2019')

                                        ax1.barh(x_axis, y_axis, color='g')
                                        for i, v in enumerate(y_axis):
                                            ax1.text(v, i - .25, str(v), color='g', fontweight='bold')
                                        ax1.set_yticklabels(x_axis, minor=False)
                                        ax1.set_title('daily quantity sold by brand')

                                        patches, texts, enumerate = ax2.pie(sizes, startangle=90, autopct='%2.f%%')
                                        ax2.legend(patches, labels, loc="best", title='gali shoes')

                                        ax2.set_title('daily gross profit by brand')
                                        ax2.axis('equal')
                                        ax2.axis('off')
                                        ax2.text(-0.1, 0,"total daily gross profit: "+str(Profits_b_sum) + '$', size=12, ha="center",
                                                 transform=ax2.transAxes)
                                        ax2.text(-0.08, -0.2,"total daily net income: "+str(net_income) + '$', size=12, ha="center",
                                                 transform=ax2.transAxes)
                                        ax2.text(0.05, -0.4,"daily income after pay to employee: "+str(after_pay_employee) + '$', size=12, ha="center",
                                                 transform=ax2.transAxes)
                                        plt.tight_layout()
                                        plt.show()

                                    elif len(ent_text.get()) == 0 :
                                        lbl_day = Label(root1, text="                  Please enter a day                     ", fg="red")
                                        lbl_day.grid(row=30, column=4)

                                    else:
                                        lbl_data = Label(root1, text="The date is not registered in the database")
                                        lbl_data.grid(column=4, row=30)
                        except:
                            print()
            #החלון של הדוח היומי אשר מופיע בו חודש ויום ולחיצת כפתור שלוחצים על search לוקח אותך לפונקציה chek_date
            root1 = Tk()
            root1.minsize(400, 200)
            root1.title("daily report 2019")

            combo1 = Combobox(root1, values=['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August',
                                            'September', 'October', 'November', 'December'])
            combo1.current(0)
            lbl1 = Label(root1, text='month:')
            combo1.grid(column=4, row=0)
            lbl1.grid(column=3, row=0)
            lbl1_1 = Label(root1, text="type single number in\n the currently month (day):")
            lbl1_1.grid(column=3, row=3)
            ent_text = Entry(root1, width=23)
            ent_text.grid(column=4, row=3)

            btn = Button(root1, text="search", command=lambda: chek_date())
            btn.grid(column=4, row=6)
            root1.mainloop()

            # התניות בעת כניסה לדוח החודשי
        elif chek_btn1.get() == 2 and text1.get() != df.columns[0]:
            get_name = Label(window, text="  wrong shop try again    ", fg="black")
            get_name.grid(column=1, row=25)

        elif chek_btn1.get() == 2 and text1.get() == df.columns[0]:
            window.destroy()
            #פונקציה מציגה דוח חודשי
             #לאחר שנלחץ על הכפתור לפי נתונים קיימים בקובץ יופיע הייצוג הגרפי החודשי
            def check_co():
                global enumerate
                df1 = 'report_2019.xlsx'

                files = glob.iglob(df1)
                for fileName in files:
                    wb = load_workbook(fileName)
                    w = wb.sheetnames
                    for sheetname in w:
                        try:
                            if sheetname != combo2.get():
                                get_name = Label(root2, text="please select month from the menu", fg="red")
                                get_name.grid(column=0, row=25)

                            else:
                                root2.destroy()

                                df.set_index('gali', inplace=True)


                                fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=1)
                                fig.canvas.set_window_title('gali shoes')

                                c = df["costumer price"]
                                p = df["purchase"]
                                prices = c + p
                                y_axis = df.loc[:][:].sum(axis=1, skipna=True) - prices
                                x_axis = df.index


                                ax1.barh(x_axis, y_axis, color='g')
                                for i, v in enumerate(y_axis):
                                    ax1.text(v + 3, i - .25, str(v), color='g', fontweight='bold')
                                ax1.set_yticklabels(x_axis, minor=False)
                                ax1.set_title('monthly quantity sold by brand')


                                labels = df.index
                                c1 = df["costumer price"]
                                z1 = df["purchase"]
                                prices = c1 + z1
                                sum_brand_month = df.loc[:][:].sum(axis=1, skipna=True) - prices
                                print(sum_brand_month)
                                Revenue = sum_brand_month * c1
                                Expenses = sum_brand_month * z1
                                Employee_salary=400*26

                                monthly_profit_b = Revenue - Expenses
                                tax = 0.10
                                monthly_profit_b_sum=monthly_profit_b.sum(axis=0, skipna=True)
                                net_income1 = monthly_profit_b.sum(axis=0, skipna=True) - (monthly_profit_b.sum(axis=0, skipna=True) * tax)
                                after_pay_employee1 = monthly_profit_b.sum(axis=0, skipna=True) - (
                                            monthly_profit_b.sum(axis=0, skipna=True) * tax) - Employee_salary
                                sizes = monthly_profit_b


                                patches, texts, enumerate = ax2.pie(sizes, startangle=90, autopct='%2.f%%')
                                ax2.legend(patches, labels, loc="best", title='gali shoes')
                                ax2.set_title('monthly gross profit by brand')
                                fig.canvas.set_window_title('monthly report 2019')

                                ax2.axis('equal')
                                ax2.text(-0.07, 0, "total monthly gross profit: " + str(monthly_profit_b_sum) + '$', size=12, ha="center",
                                         transform=ax2.transAxes)
                                ax2.text(-0.05, -0.2, "total monthly net income: " + str(net_income1) + '$', size=12, ha="center",
                                         transform=ax2.transAxes)
                                ax2.text(0.085, -0.4, "monthly income after pay to employee: " + str(after_pay_employee1) + '$', size=12, ha="center",
                                         transform=ax2.transAxes)
                                plt.tight_layout()
                                plt.show()

                        except:
                            print()

            # החלון של הדוח החודשי אשר מופיע בו חודש שגוללים ובוחרים את החודש הרצוי ולחיצת כפתור שלוחצים על search לוקח אותך לפונקציה chek_dat
            root2 = Tk()
            root2.title("monthly report 2019")
            root2.minsize(400, 200)

            combo2 = Combobox(root2, values=['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August',
                                            'September', 'October', 'November', 'December'])
            combo2.current(0)
            lbl2 = Label(root2, text='your Choise: ?')
            btn2 = Button(root2, text="search", command=check_co)

            combo2.grid(column=0, row=0)
            lbl2.grid(column=0, row=2)
            btn2.grid(column=0, row=1)

            root2.mainloop()

#פותח חלון שבו כפתורים שתוכל לבחור לבחירתך דוח יומי או
# דוח חודשי ושם החנות של אותו עסק "gali" ולאחר מכן לחיצה על
# כפתור send תיקח אותנו להמשך הקוד דרך הפונקציה check_val
window = Tk()
window.title("report app")
window.minsize(300, 180)

lbl = Label(window, text="shop name:")
text1 = Entry(window)
lbl.grid(column=0, row=0)
text1.grid(column=1, row=0)

chek_btn1 = IntVar()

r1 = Radiobutton(window, text="daily report      ", var=chek_btn1, value=1)
r2 = Radiobutton(window, text="monthly report", var=chek_btn1, value=2)

r1.grid(row=1, column=1)
r2.grid(row=3, column=1)

send_btn = Button(window, text="Send", command=lambda: check_val())
send_btn.grid(column=1, row=20)

window.mainloop()
